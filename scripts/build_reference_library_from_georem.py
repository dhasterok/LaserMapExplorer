"""Builds resources/calibration/reference_materials/<STANDARD>.yaml files from
a downloaded GeoReM "preferred values" workbook (one sheet per reference
material, exported from https://georem.earth or its predecessor).

Usage:
    python scripts/build_reference_library_from_georem.py [xlsx_path]

Defaults to resources/calibration/reference_materials/georem.xlsx. Each
sheet has a fixed shape: title/citation/date in rows 1-3, a header row
(Item, less/more, Value, Uncertainty, Uncertainty Type, Unit, GeoReM-Id,
Comment) in row 4, then one data row per element, oxide, or isotope ratio.
Rows with unit '%m/m' (oxide wt%) or 'µg/g' (elemental ppm) are concentration
data (-> ReferenceMaterial.analytes); rows with a blank unit whose Item
matches a plain "NNNSymbol/NNNSymbol" atom-ratio pattern (e.g. "206Pb/204Pb")
are certified isotope-ratio data (-> ReferenceMaterial.isotope_ratios, see
reflib.py). Delta values, permil units, and annotated ratios (e.g. activity
ratios like "230Th/232Th(act)") are still skipped.

Oxide rows are converted to elemental ppm via stoichiometry (using the same
atomic-weight table other parts of this repo already use --
global_geochemistry's MolecularWeightCalculator), since GeoReM reports major
elements as oxide wt% (Na2O, Al2O3, SiO2, CaO, ...) while this project's raw
LA-ICP-MS files report elemental CPS. GeoReM's concentrations are elemental
(isotope-independent), but this project's YAML schema keys analytes by
"Element+Mass" to match raw-file column names 1:1 (see reflib.py) -- so each
element is assigned the isotope mass this project's raw files actually
monitor (KNOWN_RAW_ISOTOPE_MASSES), falling back to that element's most
abundant natural isotope (from resources/app_data/isotope_info.csv) for
elements outside that set, so the library stays useful for future standards/
methods that monitor different isotopes.

Output overwrites any existing YAML for the same standard name -- rerun this
whenever a new georem.xlsx is downloaded to refresh the library.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

import openpyxl
import pandas as pd
from global_geochemistry.utils.molecular import MolecularWeightCalculator

from src.calibration.reflib import ReferenceAnalyte, ReferenceIsotopeRatio, ReferenceMaterial, save_reference_material

DEFAULT_XLSX_PATH = project_root / "resources" / "calibration" / "reference_materials" / "georem.xlsx"
OUTPUT_DIR = project_root / "resources" / "calibration" / "reference_materials"
ISOTOPE_TABLE_PATH = project_root / "resources" / "app_data" / "isotope_info.csv"

# GeoReM sheet name -> the standard label actually used in this project's raw
# filenames. Per the "NISTSRM610 is NIST610" convention: the NIST glasses
# drop "SRM" in the filenames this pipeline parses (e.g. "NIST610 - 1.csv").
SHEET_NAME_OVERRIDES = {
    "NISTSRM610": "NIST610",
    "NISTSRM612": "NIST612",
    "NISTSRM614": "NIST614",
}

# Per-sheet layout overrides for sheets that don't match the standard shape
# documented in this module's docstring (title/citation/date in rows 1-3,
# 8-column header in row 4, data from row 5). Only WC-1 differs so far: its
# citation and date are combined into a single row 2 (no separate row 3),
# so its header is one row earlier, and it has no "Uncertainty Type" column
# at all (only 7 columns) -- every value on that sheet is already known to
# be reported as 1SE (the user converted GeoReM's originally-reported 2SD,
# n=100, to 1SE by hand before this workbook was saved; the original 2SD
# value is preserved per-row in the Comment column for traceability, which
# already flows into the generated YAML's ``source`` field unchanged).
SHEET_LAYOUT_OVERRIDES = {
    "WC-1": {
        "metadata_rows": 2, "header_row": 3, "data_start_row": 4,
        "has_uncertainty_type_column": False, "default_uncertainty_type": "SEM",
    },
}

# The isotope this project's raw files actually monitor per element (see the
# "Time [Sec],Li7,Na23,Mg24,..." raw-file column header). GeoReM's values are
# elemental, so a YAML key must pick *some* mass number to match those
# columns 1:1; anything not listed here falls back to the most-abundant
# natural isotope (see `_fallback_isotope_masses`).
KNOWN_RAW_ISOTOPE_MASSES = {
    "Li": 7, "Na": 23, "Mg": 24, "Al": 27, "Si": 29, "P": 31, "Ca": 43,
    "Ti": 49, "V": 51, "Cr": 53, "Mn": 55, "Fe": 57, "Sr": 88, "Y": 89,
    "Zr": 90, "La": 139, "Ce": 140, "Pr": 141, "Nd": 146, "Sm": 147,
    "Eu": 153, "Gd": 157, "Tb": 159, "Dy": 163, "Ho": 165, "Er": 166,
    "Tm": 169, "Yb": 172, "Lu": 175, "Hf": 178, "Th": 232, "U": 238,
}

# Concentration units only -- isotope ratios/deltas use permil units
# (e.g. '‰NIST976') and are not concentration data.
_CONCENTRATION_UNITS = {"%m/m", "µg/g"}

_OXIDE_RE = re.compile(r"^([A-Z][a-z]?)(\d*)O(\d*)$")
_ELEMENT_RE = re.compile(r"^[A-Z][a-z]?$")
_UNCERTAINTY_TYPE_MAP = {"95%CL": "95CI"}

# GeoREM isotope-ratio items are reported mass-first (e.g. "206Pb/204Pb"),
# the reverse of this project's element-first analyte-column convention
# ("Pb206") -- matched against the RAW item string (before _strip_suffix),
# since a "(...)" suffix on a ratio item (e.g. "230Th/232Th(act)", an
# *activity* ratio) is a physically different quantity from the plain atom
# ratio "230Th/232Th", unlike an oxide's cosmetic "(t)" suffix.
_RATIO_RE = re.compile(r"^(\d+)([A-Za-z]{1,2})/(\d+)([A-Za-z]{1,2})$")


def _fallback_isotope_masses(path: Path = ISOTOPE_TABLE_PATH) -> dict[str, int]:
    """Most-abundant natural isotope mass per element, for elements outside
    KNOWN_RAW_ISOTOPE_MASSES."""
    table = pd.read_csv(path)
    masses = {}
    for symbol, group in table.groupby("symbol"):
        best = group.loc[group["abundance_nominal"].idxmax()]
        masses[symbol] = int(best["atomic_mass"])
    return masses


def _isotope_mass_for(element: str, fallback: dict[str, int]) -> int | None:
    return KNOWN_RAW_ISOTOPE_MASSES.get(element) or fallback.get(element)


def _strip_suffix(item: str) -> str:
    """'Fe2O3(t)' -> 'Fe2O3' -- '(t)' means "total (Fe as this species)" and
    doesn't change how the formula itself is parsed."""
    return re.sub(r"\(.*\)$", "", item).strip()


def _parse_oxide(formula: str) -> tuple[str, int, int] | None:
    m = _OXIDE_RE.match(formula)
    if not m:
        return None
    cation, n_cation_str, n_oxygen_str = m.groups()
    return cation, int(n_cation_str or 1), int(n_oxygen_str or 1)


def _oxide_to_element_ppm(
    mwc: MolecularWeightCalculator, oxide_formula: str, wt_pct: float, unc_wt_pct: float | None
) -> tuple[str, float, float | None] | None:
    """(element_symbol, element_ppm, element_uncertainty_ppm), or None if the
    formula can't be parsed/has no known atomic weight."""
    parsed = _parse_oxide(oxide_formula)
    if parsed is None:
        return None
    element, n_cation, _n_oxygen = parsed
    element_mass = mwc.element_data.get(element)
    if element_mass is None:
        return None
    oxide_mass = mwc.molecular_weight(oxide_formula)
    fraction = (n_cation * element_mass) / oxide_mass
    element_ppm = wt_pct * 10000.0 * fraction
    element_unc_ppm = unc_wt_pct * 10000.0 * fraction if unc_wt_pct is not None else None
    return element, element_ppm, element_unc_ppm


def _map_uncertainty_type(raw_type) -> tuple[str, str]:
    """(schema_uncertainty_type, note) -- note is non-empty only when the
    original type was ambiguous/unrecognized, so a fabricated '1SD' default
    isn't silently presented as a real reported 1SD figure."""
    if raw_type is None:
        return "1SD", ""
    key = str(raw_type).strip()
    if key in _UNCERTAINTY_TYPE_MAP:
        return _UNCERTAINTY_TYPE_MAP[key], ""
    if key.lower() in ("sd", "1sd"):
        return "1SD", ""
    if key.lower() == "2sd":
        return "2SD", ""
    if key.lower() == "2se":
        return "2SE", ""
    return "1SD", f" (original uncertainty type '{raw_type}' unrecognized -- treated as 1SD)"


def _parse_ratio_item(item: str) -> tuple[int, str, int, str] | None:
    """(numerator_mass, numerator_element, denominator_mass, denominator_element),
    or None if ``item`` isn't a plain "NNNSymbol/NNNSymbol" atom ratio.

    Matched against the RAW item string -- deliberately does NOT strip a
    "(...)" suffix first, unlike ``_strip_suffix``'s use for oxide/element
    items. "230Th/232Th(act)" (an *activity* ratio) is a different
    physical quantity from the plain atom ratio "230Th/232Th", so
    collapsing them via suffix-stripping would silently conflate two
    distinct reported values.
    """
    m = _RATIO_RE.match(item)
    if not m:
        return None
    num_mass, num_el, den_mass, den_el = m.groups()
    return int(num_mass), num_el, int(den_mass), den_el


def _convert_ratio_row(
    item: str, value, uncertainty, uncertainty_type, base_source: str, georem_id, comment, sheet_name: str,
) -> tuple[str, ReferenceIsotopeRatio] | None:
    """(schema key e.g. "Pb206/Pb204", ReferenceIsotopeRatio), or None if
    ``item`` doesn't parse as a plain isotope-ratio item."""
    parsed = _parse_ratio_item(item)
    if parsed is None:
        return None
    num_mass, num_el, den_mass, den_el = parsed
    key = f"{num_el}{num_mass}/{den_el}{den_mass}"

    unc_type_str = str(uncertainty_type).strip() if uncertainty_type is not None else ""
    ratio_uncertainty = float(uncertainty) if uncertainty is not None else None
    source_parts = [base_source]
    if georem_id:
        source_parts.append(str(georem_id))
    if comment:
        source_parts.append(str(comment))

    if "%" in unc_type_str:
        # Relative-percent uncertainty (e.g. BHVO-2G's '2RSE(%)') -- don't
        # coerce into the absolute-uncertainty schema (a relative number
        # stuffed into an absolute field would be a silent unit error).
        schema_unc_type = None
        ratio_uncertainty = None
        source_parts.append(f"original uncertainty reported as {uncertainty_type} (relative) -- not converted, needs manual handling")
    elif not unc_type_str:
        # Genuinely missing from the source (e.g. GeoREM's NIST612/614
        # isotope-ratio rows) -- surfaced explicitly, not guessed at.
        schema_unc_type = None
        print(
            f"  WARNING: {sheet_name} '{item}' has no reported Uncertainty Type -- "
            "stored as uncertainty_type: null, needs user verification (NOT assumed "
            "to match another sheet's/standard's sourcing)."
        )
    else:
        schema_unc_type, unc_note = _map_uncertainty_type(uncertainty_type)
        if unc_note:
            source_parts.append(unc_note.strip(" ()"))

    entry = ReferenceIsotopeRatio(
        numerator_element=num_el, numerator_mass=num_mass,
        denominator_element=den_el, denominator_mass=den_mass,
        value=float(value), uncertainty=ratio_uncertainty, uncertainty_type=schema_unc_type,
        source=", ".join(p for p in source_parts if p),
    )
    return key, entry


def convert_sheet(
    ws, sheet_name: str, mwc: MolecularWeightCalculator, fallback_masses: dict[str, int]
) -> tuple[ReferenceMaterial, list]:
    layout = SHEET_LAYOUT_OVERRIDES.get(sheet_name, {})
    metadata_rows = layout.get("metadata_rows", 3)
    data_start_row = layout.get("data_start_row", 5)
    has_uncertainty_type_column = layout.get("has_uncertainty_type_column", True)
    default_uncertainty_type = layout.get("default_uncertainty_type")

    title = ws.cell(row=1, column=1).value or ""
    if metadata_rows >= 3:
        citation = ws.cell(row=2, column=1).value or ""
        date = ws.cell(row=3, column=1).value or ""
        base_source = f"{citation} {date}".strip()
    else:
        # Citation and date combined into one row (e.g. WC-1).
        base_source = str(ws.cell(row=2, column=1).value or "").strip()

    analytes: dict[str, ReferenceAnalyte] = {}
    isotope_ratios: dict[str, ReferenceIsotopeRatio] = {}
    skipped = []
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if has_uncertainty_type_column:
            item, _less_more, value, uncertainty, uncertainty_type, unit, georem_id, comment = row[:8]
        else:
            item, _less_more, value, uncertainty, unit, georem_id, comment = row[:7]
            uncertainty_type = None
        if item is None or value is None:
            continue
        # Some sheets pad the unit cell with trailing whitespace/NBSP
        # (observed on WC-1: 'µg/g\xa0\xa0\xa0') -- normalize before the
        # exact-match check below, or every row on that sheet silently
        # gets skipped as "not concentration data".
        unit_clean = str(unit).strip() if unit is not None else None
        if unit_clean not in _CONCENTRATION_UNITS:
            # Blank/non-concentration unit -- not necessarily junk, try
            # isotope-ratio parsing before giving up on the row entirely
            # (previously every such row, including real GeoREM ratio
            # data, was silently dropped here with no trace).
            ratio_entry = _convert_ratio_row(
                str(item), value, uncertainty, uncertainty_type, base_source, georem_id, comment, sheet_name,
            )
            if ratio_entry is not None:
                key, entry = ratio_entry
                isotope_ratios[key] = entry
            else:
                note = " (annotated/activity ratio, not a plain atom ratio -- not imported)" if "(" in str(item) else ""
                skipped.append(f"{item}{note}")
            continue
        unit = unit_clean

        item_clean = _strip_suffix(str(item))
        if uncertainty_type is not None:
            schema_unc_type, unc_note = _map_uncertainty_type(uncertainty_type)
        elif default_uncertainty_type is not None:
            # Known with certainty from the sheet's own layout/convention
            # (not a guess), so no "unrecognized" note -- see
            # SHEET_LAYOUT_OVERRIDES's docstring for why WC-1 uses this.
            schema_unc_type, unc_note = default_uncertainty_type, ""
        else:
            schema_unc_type, unc_note = _map_uncertainty_type(None)
        source_note = ""

        if unit == "%m/m":
            converted = _oxide_to_element_ppm(
                mwc, item_clean, float(value), float(uncertainty) if uncertainty is not None else None
            )
            if converted is None:
                skipped.append(item)
                continue
            element, ppm_value, ppm_uncertainty = converted
            source_note = f"converted from {item} ({value}% m/m)"
        elif _ELEMENT_RE.match(item_clean):
            element = item_clean
            ppm_value = float(value)
            ppm_uncertainty = float(uncertainty) if uncertainty is not None else None
        else:
            skipped.append(item)
            continue

        mass = _isotope_mass_for(element, fallback_masses)
        if mass is None:
            skipped.append(item)
            continue

        source_parts = [base_source]
        if georem_id:
            source_parts.append(str(georem_id))
        if comment:
            source_parts.append(str(comment))
        if source_note:
            source_parts.append(source_note)
        source = ", ".join(p for p in source_parts if p) + unc_note

        analytes[f"{element}{mass}"] = ReferenceAnalyte(
            element=element, mass=mass, value=ppm_value, uncertainty=ppm_uncertainty,
            uncertainty_type=schema_unc_type, units="ppm", source=source,
        )

    standard = SHEET_NAME_OVERRIDES.get(sheet_name, sheet_name)
    material = ReferenceMaterial(
        standard=standard, description=str(title), source=base_source,
        analytes=analytes, isotope_ratios=isotope_ratios,
    )
    return material, skipped


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX_PATH
    if not xlsx_path.exists():
        print(f"Workbook not found: {xlsx_path}")
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    mwc = MolecularWeightCalculator()
    fallback_masses = _fallback_isotope_masses()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for sheet_name in wb.sheetnames:
        material, skipped = convert_sheet(wb[sheet_name], sheet_name, mwc, fallback_masses)
        out_path = OUTPUT_DIR / f"{material.standard}.yaml"
        save_reference_material(material, out_path)
        print(f"Wrote {out_path} ({len(material.analytes)} analytes, {len(material.isotope_ratios)} isotope ratios)")
        if skipped:
            print(f"  skipped {len(skipped)} row(s) with no elemental match: {skipped}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
