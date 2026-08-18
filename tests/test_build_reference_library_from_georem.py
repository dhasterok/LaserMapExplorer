"""Regression test for the georem.xlsx -> YAML conversion script's WC-1
layout special-case (see SHEET_LAYOUT_OVERRIDES in the script) -- WC-1's
sheet has a different shape than every other sheet the script otherwise
assumes (2 metadata rows instead of 3, no "Uncertainty Type" column), and
running the script unmodified over it used to crash.

Pure Python -- no PyQt/QApplication needed.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from global_geochemistry.utils.molecular import MolecularWeightCalculator

from scripts.build_reference_library_from_georem import convert_sheet


def _wc1_shaped_workbook():
    """A minimal synthetic workbook matching WC-1's real layout: title row,
    citation+date combined into one row, a 7-column header (no Uncertainty
    Type), data from row 4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WC-1"
    ws.append(["WC-1:  marine calcite cement"])
    ws.append(["Roberts et al. (2017), doi:10.1002/2016GC006784."])
    ws.append(["Item", "less/more", "Value", "Uncertainty", "Unit", "GeoReM-Id", "Comment"])
    ws.append(["Mg", None, 663, 38.5, "µg/g\xa0\xa0\xa0", "GeoReM 10551", "Uncertainty in GEOREM is reported as 2SD, 770 µg/g"])
    ws.append(["Sr", None, 1500, 95, "µg/g\xa0\xa0\xa0", None, "Uncertainty in GEOREM is reported as 2SD, 1900 µg/g"])
    return ws


def test_convert_sheet_parses_wc1_layout_without_raising():
    ws = _wc1_shaped_workbook()
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "WC-1", mwc, fallback_masses={})

    assert material.standard == "WC-1"
    assert not skipped
    assert "Mg24" in material.analytes
    assert "Sr88" in material.analytes


def test_convert_sheet_wc1_uncertainty_type_is_sem_not_unrecognized():
    ws = _wc1_shaped_workbook()
    mwc = MolecularWeightCalculator()
    material, _ = convert_sheet(ws, "WC-1", mwc, fallback_masses={})

    analyte = material.analytes["Mg24"]
    assert analyte.uncertainty_type == "SEM"
    assert "unrecognized" not in analyte.source


def test_convert_sheet_wc1_source_preserves_original_2sd_comment():
    ws = _wc1_shaped_workbook()
    mwc = MolecularWeightCalculator()
    material, _ = convert_sheet(ws, "WC-1", mwc, fallback_masses={})

    assert "2SD, 770" in material.analytes["Mg24"].source


def test_convert_sheet_normal_sheet_layout_unaffected():
    """Sanity check: a sheet NOT in SHEET_LAYOUT_OVERRIDES still uses the
    original 3-metadata-row/8-column/row-5-data layout unchanged."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TESTSTD"
    ws.append(["Test Standard"])
    ws.append(["Some Citation"])
    ws.append(["2020"])
    ws.append(["Item", "less/more", "Value", "Uncertainty", "Uncertainty Type", "Unit", "GeoReM-Id", "Comment"])
    ws.append(["Sr", None, 400, 5, "1SD", "µg/g", None, None])

    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})
    assert not skipped
    assert material.analytes["Sr88"].uncertainty_type == "1SD"
    assert material.analytes["Sr88"].value == pytest.approx(400.0)


def _normal_sheet_with_rows(*data_rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TESTSTD"
    ws.append(["Test Standard"])
    ws.append(["Some Citation"])
    ws.append(["2020"])
    ws.append(["Item", "less/more", "Value", "Uncertainty", "Uncertainty Type", "Unit", "GeoReM-Id", "Comment"])
    for row in data_rows:
        ws.append(row)
    return ws


def test_convert_sheet_parses_plain_isotope_ratio_row():
    ws = _normal_sheet_with_rows(
        ["206Pb/204Pb", None, 17.047, 0.0018, "2SD", None, "GeoReM 5211", "mean value"],
    )
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})

    assert not skipped
    assert "Pb206/Pb204" in material.isotope_ratios
    entry = material.isotope_ratios["Pb206/Pb204"]
    assert entry.numerator == "Pb206"
    assert entry.denominator == "Pb204"
    assert entry.value == pytest.approx(17.047)
    assert entry.uncertainty_type == "2SD"
    assert entry.uncertainty == pytest.approx(0.0018)
    assert "GeoReM 5211" in entry.source
    # Must not also be treated as elemental/analyte data.
    assert not material.analytes


def test_convert_sheet_ratio_row_blank_uncertainty_type_is_none_with_warning(capsys):
    # Matches real GeoREM data: NIST612/614's isotope-ratio rows report no
    # Uncertainty Type at all.
    ws = _normal_sheet_with_rows(
        ["206Pb/204Pb", None, 17.094, 0.0026, None, None, None, None],
    )
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})

    assert not skipped
    entry = material.isotope_ratios["Pb206/Pb204"]
    assert entry.uncertainty_type is None
    assert entry.value == pytest.approx(17.094)
    captured = capsys.readouterr()
    assert "WARNING" in captured.out
    assert "no reported Uncertainty Type" in captured.out


def test_convert_sheet_activity_ratio_suffix_not_parsed_as_ratio():
    ws = _normal_sheet_with_rows(
        ["230Th/232Th(act)", None, 0.876, 0.02, "SD", None, None, None],
    )
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})

    assert not material.isotope_ratios
    assert not material.analytes
    assert len(skipped) == 1
    assert "230Th/232Th(act)" in skipped[0]
    assert "activity ratio" in skipped[0]


def test_convert_sheet_ratio_row_2se_maps_correctly():
    ws = _normal_sheet_with_rows(
        ["143Nd/144Nd", None, 0.511927, 4e-06, "2SE", None, None, None],
    )
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})

    assert not skipped
    entry = material.isotope_ratios["Nd143/Nd144"]
    assert entry.uncertainty_type == "2SE"
    assert entry.uncertainty_1sd() == pytest.approx(2e-06)


def test_convert_sheet_ratio_row_percent_relative_uncertainty_not_coerced():
    ws = _normal_sheet_with_rows(
        ["230Th/232Th", None, 4.737e-06, 9.3e-09, "2RSE(%)", None, None, None],
    )
    mwc = MolecularWeightCalculator()
    material, skipped = convert_sheet(ws, "TESTSTD", mwc, fallback_masses={})

    assert not skipped
    entry = material.isotope_ratios["Th230/Th232"]
    assert entry.value == pytest.approx(4.737e-06)
    assert entry.uncertainty is None
    assert entry.uncertainty_type is None
    assert "2RSE" in entry.source
    assert "not converted" in entry.source
