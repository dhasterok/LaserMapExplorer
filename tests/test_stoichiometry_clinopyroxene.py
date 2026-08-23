"""Consistency test for clinopyroxene.yaml -- not a fresh workbook
cross-check, since this config's *site allocation* is deliberately the same
Ca-Mg-Fe quadrilateral engine as ``resources/minerals/pyroxene.yaml`` under
a clinopyroxene-specific name (see that file's comment and
``test_stoichiometry_pyroxene_workbook_reference.py``, which already
validates this exact engine cell-by-cell against a real EPMA clinopyroxene
analysis). Its *end-member* step now intentionally diverges: clinopyroxene
uses ``pyroxene_quad_jd_ae`` (adds the jadeite/aegirine Na-pyroxene branch,
see endmembers.py), while pyroxene.yaml still uses the plain quad-only
``pyroxene_quad``. This test confirms apfu still matches pyroxene.yaml
exactly (site allocation is unaffected by the end-member method choice),
and that the Wo/En/Fs values agree too for this reference composition,
which has no Na2O -- so jadeite/aegirine both come out at 0% and the quad
math reduces to the same numbers pyroxene.yaml reports. Duplicating the
full workbook parse here would just re-test the same site-allocation code
path a second time.

Pure Python -- no PyQt/QApplication needed.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

PYROXENE_YAML_PATH = project_root / "resources" / "minerals" / "pyroxene.yaml"
CLINOPYROXENE_YAML_PATH = project_root / "resources" / "minerals" / "clinopyroxene.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"


@pytest.fixture(scope="module")
def reference_analysis():
    ws = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)["Pyroxene"]
    analysis = {}
    for row in range(11, 21):
        oxide = ws.cell(row=row, column=1).value
        if oxide is None or oxide == "Fe2O3":
            continue
        wt_pct = ws.cell(row=row, column=2).value
        if wt_pct is None:
            continue
        analysis[oxide] = float(wt_pct)
    return analysis


def test_clinopyroxene_reproduces_pyroxene_apfu_and_end_members(reference_analysis):
    pyroxene_config = load_mineral_config(PYROXENE_YAML_PATH)
    clinopyroxene_config = load_mineral_config(CLINOPYROXENE_YAML_PATH)

    r_px = pipeline.calculate(reference_analysis, pyroxene_config, input_mode="wt_percent", redox_method="droop_1987")
    r_cpx = pipeline.calculate(reference_analysis, clinopyroxene_config, input_mode="wt_percent", redox_method="droop_1987")

    assert r_cpx.apfu == r_px.apfu
    # No Na2O in this reference analysis -> jadeite/aegirine both 0%, and the
    # shared Wo/En/Fs quad values should agree exactly with pyroxene.yaml's.
    assert r_cpx.end_members["jadeite"] == 0.0
    assert r_cpx.end_members["aegirine"] == 0.0
    for member in ("wollastonite", "enstatite", "ferrosilite"):
        assert r_cpx.end_members[member] == pytest.approx(r_px.end_members[member])
    assert r_cpx.end_members["enstatite"] > r_cpx.end_members["ferrosilite"] > r_cpx.end_members["wollastonite"]


def test_clinopyroxene_abbreviation_is_distinct_from_pyroxene():
    pyroxene_config = load_mineral_config(PYROXENE_YAML_PATH)
    clinopyroxene_config = load_mineral_config(CLINOPYROXENE_YAML_PATH)
    assert clinopyroxene_config.mineral == "clinopyroxene"
    assert clinopyroxene_config.abbreviation == "Cpx"
    assert clinopyroxene_config.abbreviation != pyroxene_config.abbreviation
