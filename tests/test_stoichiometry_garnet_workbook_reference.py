"""Cross-checks the stoichiometry pipeline against ``tests/normalization.v17.xlsx``'s
'Garnet' sheet -- a hand-built, cell-by-cell recalculation of a real EPMA garnet
analysis (Napak, Uganda) that solves for Fe3+ by charge balance.

Pure Python -- no PyQt/QApplication needed.

This test file originally characterized two bugs found by this comparison:
apfu were reported on the wrong normalization basis (oxygen-fixed S instead of
cation-fixed T, ~4.4% inflation on this sample), and site allocation silently
dropped elements once a site's target was reached (Mn and Ti vanished
entirely here). Both are now fixed in ``redox.py``/``sites.py``/``normalize.py``,
cross-checked against a second independent reference, MinPlotX
(``../MinPlotX/``'s ``garnet_Fe3unknown.m``) -- see that file for how each
site apportions elements it shares with a neighboring site.

Everything below now matches the workbook to within ~0.1%, *except* two small,
deliberate, documented divergences:

1. **Z-site (tetrahedral) total is 3.000, not the workbook's 2.996.** This
   repo's ``garnet.yaml`` models Al spilling into Z when Si is slightly
   deficient (``priority: [Si, Al, Fe3]  # fill order when site is
   under-occupied``) -- MinPlotX does the same (`garnet_Fe3unknown.m`'s
   `Al(T)` block). The workbook's simpler cell-by-cell model doesn't spill Al
   into Z at all. Since our own config already documented this as intentional
   before this test existed, MinPlotX (not the workbook) is treated as
   authoritative here.
2. **Y-site (octahedral) total is ~0.2% low.** MinPlotX's garnet model also
   lets Mg/Fe2+/Mn spill into Y if Al/Cr/Fe3+/Ti don't fill it -- not modeled
   in ``garnet.yaml`` (see the comment on its Y site). Deferred rather than
   fixed here: it needs new elements-list semantics (an element belonging to
   a site only as a last-resort fallback), and the effect is small.

Point 1 also propagates into the end-member split (extra Al ends up in Z
instead of Y, shifting the Y-site's Al:Fe3+ ratio slightly), which is why
grossular/andradite are within ~1% rather than ~0.1%.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

GARNET_YAML_PATH = project_root / "resources" / "minerals" / "garnet.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"

# Oxide input key (as it appears in the workbook's 'Analysis' block, A11:B19)
# -> cation species key used in this codebase's apfu dicts.
_OXIDE_TO_CATION = {
    "SiO2": "Si", "TiO2": "Ti", "Al2O3": "Al", "Cr2O3": "Cr",
    "FeO": "Fe2", "MnO": "Mn", "MgO": "Mg", "CaO": "Ca",
}
# Workbook site label (B/C/D59) -> this config's site name.
_WORKBOOK_SITE_TO_CONFIG_SITE = {"cubic": "X", "octahedral": "Y", "tetrahedral": "Z"}


@pytest.fixture(scope="module")
def garnet_config():
    return load_mineral_config(GARNET_YAML_PATH)


@pytest.fixture(scope="module")
def workbook_garnet_sheet():
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    return wb["Garnet"]


@pytest.fixture(scope="module")
def workbook_reference(workbook_garnet_sheet):
    """Pull the input analysis and every expected output straight from the
    workbook's cells (rather than hand-transcribing), so this stays tied to
    the committed source-of-truth file.
    """
    ws = workbook_garnet_sheet

    analysis = {}
    expected_apfu = {}
    for row in range(11, 20):
        oxide = ws.cell(row=row, column=1).value  # col A
        if oxide == "Fe2O3":
            continue  # not measured -- Fe3+ is solved for by charge balance
        wt_pct = ws.cell(row=row, column=2).value  # col B: analysis wt%
        final_apfu = ws.cell(row=row, column=7).value  # col G: atom units (final)
        analysis[oxide] = float(wt_pct)
        expected_apfu[_OXIDE_TO_CATION[oxide]] = float(final_apfu)
    expected_apfu["Fe3"] = float(ws["G15"].value)  # Fe2O3 row's atom-units column

    expected_sites = {
        _WORKBOOK_SITE_TO_CONFIG_SITE[label]: float(ws.cell(row=59, column=col).value)
        for label, col in (("cubic", 2), ("octahedral", 3), ("tetrahedral", 4))
    }

    # End-member column F, "mol % with Fe3+, etc." -- the only one of the
    # workbook's three end-member columns that accounts for the charge-balanced
    # Fe3+ split (the other two are simplified all-Fe schemes). 'Ca-Ti Gt'
    # (0.715%) is dropped: this config's end_members.members list doesn't
    # model it, so it has no counterpart to compare against.
    expected_end_members = {}
    for row in range(49, 56):
        name = ws.cell(row=row, column=5).value
        pct = ws.cell(row=row, column=6).value
        if name != "Ca-Ti Gt":
            expected_end_members[name] = float(pct or 0.0)

    return {
        "analysis": analysis,
        "apfu": expected_apfu,
        "sites": expected_sites,
        "end_members": expected_end_members,
        "oxide_total_pct": float(ws["B21"].value),
    }


@pytest.fixture(scope="module")
def result(garnet_config, workbook_reference):
    return pipeline.calculate(
        workbook_reference["analysis"], garnet_config,
        input_mode="wt_percent", redox_method="droop_1987",
    )


def test_oxide_total_matches_workbook(result, workbook_reference):
    assert result.oxide_total_pct == pytest.approx(workbook_reference["oxide_total_pct"], rel=1e-3)


def test_fe3_charge_balance_split_matches_workbook(result, workbook_reference):
    assert result.redox.species_3plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe3"], rel=1e-3)
    assert result.redox.species_2plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe2"], rel=1e-3)


def test_apfu_matches_workbook(result, workbook_reference):
    """Every cation's final apfu, on the same T (cation-fixed) basis the
    workbook itself uses -- no rescaling trick needed now that apfu are
    reported on the correct basis.
    """
    for cation, expected in workbook_reference["apfu"].items():
        assert result.apfu[cation] == pytest.approx(expected, rel=1e-3), cation


def test_site_totals_match_workbook_within_documented_gaps(result, workbook_reference):
    sites = result.site_allocation.sites
    wb = workbook_reference["sites"]
    assert not result.site_allocation.unallocated  # nothing silently dropped
    # X: no known divergence -- matches tightly.
    assert sites["X"].total == pytest.approx(wb["X"], rel=1e-3)
    assert sites["X"].elements["Mn"] == pytest.approx(0.0095, abs=2e-4)  # previously dropped to 0 entirely
    # Z: deliberately doesn't match -- see module docstring point 1 (Al
    # spillover, matching MinPlotX and this repo's own documented intent).
    assert sites["Z"].total == pytest.approx(3.0, abs=1e-6)
    assert sites["Z"].total > wb["Z"]
    # Y: small, documented, deferred gap -- see module docstring point 2.
    assert sites["Y"].total == pytest.approx(wb["Y"], rel=3e-3)


def test_end_members_match_workbook_within_documented_gap(result, workbook_reference):
    em = result.end_members
    wb_em = workbook_reference["end_members"]
    assert sum(em.values()) == pytest.approx(100.0, abs=1e-6)
    # Pyrope/almandine/spessartine (X-site fractions) aren't affected by the
    # Z-spillover gap -- match tightly.
    assert em["pyrope"] == pytest.approx(wb_em["pyrope"], rel=1e-3)
    assert em["almandine"] == pytest.approx(wb_em["almandine"], rel=1e-3)
    assert em["spessartine"] == pytest.approx(wb_em["spessartine"], rel=1e-2)
    # Grossular/andradite (Y-site Al:Fe3+ ratio) inherit the Z-spillover gap
    # (module docstring point 1) -- match to ~1%, not ~0.1%.
    assert em["grossular"] == pytest.approx(wb_em["grossular"], rel=1e-2)
    assert em["andradite"] == pytest.approx(wb_em["andradite"], rel=1.5e-2)
    assert em["uvarovite"] == pytest.approx(wb_em["uvarovite"], abs=1e-6)
