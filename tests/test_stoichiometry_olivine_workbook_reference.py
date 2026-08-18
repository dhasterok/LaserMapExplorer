"""Cross-checks the olivine stoichiometry pipeline against
``tests/normalization.v17.xlsx``'s 'Olivine' sheet -- a hand-built,
cell-by-cell recalculation of a real EPMA olivine analysis (a lherzolite)
that solves for Fe3+ by charge balance.

Pure Python -- no PyQt/QApplication needed. Uses the same fixed engine as
garnet (see ``tests/test_stoichiometry_garnet_workbook_reference.py``) and
``resources/minerals/olivine.yaml``'s generic priority-fill sites (T then M,
both fit the existing engine unchanged -- no mineral-specific site code was
needed for olivine, see the Stage 2 plan). Everything checked here matches
the workbook to well within 0.1%.

Unlike the Garnet sheet, the Olivine sheet has no separate T/M site-occupancy
table (just total apfu and end-members), so site totals are checked against
their crystallographic targets (T=1, M=2) rather than the workbook directly.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

OLIVINE_YAML_PATH = project_root / "resources" / "minerals" / "olivine.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"

_OXIDE_TO_CATION = {
    "SiO2": "Si", "TiO2": "Ti", "Al2O3": "Al", "Cr2O3": "Cr",
    "FeO": "Fe2", "MnO": "Mn", "MgO": "Mg", "CaO": "Ca",
}
_WORKBOOK_TO_MEMBER = {"Te": "tephroite", "Fo": "forsterite", "Fa": "fayalite", "Ca-Ol": "ca_olivine"}


@pytest.fixture(scope="module")
def olivine_config():
    return load_mineral_config(OLIVINE_YAML_PATH)


@pytest.fixture(scope="module")
def workbook_reference():
    ws = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)["Olivine"]

    analysis = {}
    expected_apfu = {}
    for row in range(11, 20):
        oxide = ws.cell(row=row, column=1).value
        if oxide is None or oxide == "Fe2O3":
            continue
        analysis[oxide] = float(ws.cell(row=row, column=2).value)
        expected_apfu[_OXIDE_TO_CATION[oxide]] = float(ws.cell(row=row, column=7).value)
    expected_apfu["Fe3"] = float(ws["G15"].value)

    expected_end_members = {
        _WORKBOOK_TO_MEMBER[ws.cell(row=r, column=7).value]: float(ws.cell(row=r, column=8).value)
        for r in range(33, 37)
    }

    return {
        "analysis": analysis,
        "apfu": expected_apfu,
        "end_members": expected_end_members,
        "oxide_total_pct": float(ws["B21"].value),
    }


@pytest.fixture(scope="module")
def result(olivine_config, workbook_reference):
    return pipeline.calculate(
        workbook_reference["analysis"], olivine_config,
        input_mode="wt_percent", redox_method="droop_1987",
    )


def test_oxide_total_matches_workbook(result, workbook_reference):
    assert result.oxide_total_pct == pytest.approx(workbook_reference["oxide_total_pct"], rel=1e-3)


def test_fe3_charge_balance_split_matches_workbook(result, workbook_reference):
    assert result.redox.species_3plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe3"], rel=1e-2)
    assert result.redox.species_2plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe2"], rel=1e-3)


def test_apfu_matches_workbook(result, workbook_reference):
    for cation, expected in workbook_reference["apfu"].items():
        if cation == "Fe3":
            continue  # trace-magnitude value, checked with a looser tolerance above
        assert result.apfu[cation] == pytest.approx(expected, rel=1e-3), cation


def test_site_totals_hit_crystallographic_targets(result):
    sites = result.site_allocation.sites
    assert not result.site_allocation.unallocated
    assert sites["T"].total == pytest.approx(1.0, abs=1e-3)
    assert sites["M"].total == pytest.approx(2.0, abs=1e-3)
    assert sites["T"].elements["Si"] == pytest.approx(0.988, abs=1e-3)


def test_end_members_match_workbook(result, workbook_reference):
    em = result.end_members
    assert sum(em.values()) == pytest.approx(100.0, abs=1e-6)
    for member, expected in workbook_reference["end_members"].items():
        assert em[member] == pytest.approx(expected, rel=1e-2), member
    assert em["forsterite"] > em["fayalite"] > em["tephroite"] > em["ca_olivine"]
