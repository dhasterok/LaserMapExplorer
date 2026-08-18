"""Cross-checks the plagioclase stoichiometry pipeline against
``tests/normalization.v17.xlsx``'s 'Plagioclase' sheet -- a hand-built,
cell-by-cell recalculation of a real EPMA labradorite analysis that solves
for Fe3+ by charge balance.

Pure Python -- no PyQt/QApplication needed. ``resources/minerals/plagioclase.yaml``
uses the generic priority-fill engine unchanged (no mineral-specific site code
needed, see the Stage 2 plan) -- everything below matches the workbook to
within ~0.1%.

Per user decision, this config includes Fe-redox estimation (``droop_1987``,
reusing the already-generic ``redox.py`` unchanged) even though MinPlotX's
shipped ``feldspar_MinPlotX.m`` doesn't do this -- the workbook does, and it's
free to add since the engine is mineral-agnostic already.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

PLAGIOCLASE_YAML_PATH = project_root / "resources" / "minerals" / "plagioclase.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"

_OXIDE_TO_CATION = {
    "SiO2": "Si", "TiO2": "Ti", "Al2O3": "Al", "Cr2O3": "Cr", "FeO": "Fe2",
    "MnO": "Mn", "MgO": "Mg", "CaO": "Ca", "BaO": "Ba", "Na2O": "Na", "K2O": "K",
}
_WORKBOOK_TO_MEMBER = {"An": "anorthite", "Ab": "albite", "Or": "orthoclase"}


@pytest.fixture(scope="module")
def plagioclase_config():
    return load_mineral_config(PLAGIOCLASE_YAML_PATH)


@pytest.fixture(scope="module")
def workbook_reference():
    ws = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)["Plagioclase"]

    analysis = {}
    expected_apfu = {}
    for row in range(11, 23):
        oxide = ws.cell(row=row, column=1).value
        if oxide is None or oxide == "Fe2O3":
            continue
        wt_pct = ws.cell(row=row, column=2).value
        if wt_pct is None:
            continue
        analysis[oxide] = float(wt_pct)
        expected_apfu[_OXIDE_TO_CATION[oxide]] = float(ws.cell(row=row, column=7).value)
    expected_apfu["Fe3"] = float(ws["G15"].value)

    expected_end_members = {
        _WORKBOOK_TO_MEMBER[ws.cell(row=r, column=7).value]: float(ws.cell(row=r, column=8).value)
        for r in range(35, 38)
    }

    return {
        "analysis": analysis,
        "apfu": expected_apfu,
        "t_site_total": float(ws["H42"].value),   # "Si+Ti+Al+Fe3="
        "ca_na_k_total": float(ws["H45"].value),  # "Ca+Na+K="
        "end_members": expected_end_members,
        "oxide_total_pct": float(ws["B24"].value),
    }


@pytest.fixture(scope="module")
def result(plagioclase_config, workbook_reference):
    return pipeline.calculate(
        workbook_reference["analysis"], plagioclase_config,
        input_mode="wt_percent", redox_method="droop_1987",
    )


def test_oxide_total_matches_workbook(result, workbook_reference):
    assert result.oxide_total_pct == pytest.approx(workbook_reference["oxide_total_pct"], rel=1e-3)


def test_fe3_charge_balance_split_matches_workbook(result, workbook_reference):
    assert result.redox.species_3plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe3"], rel=1e-2)
    assert result.redox.species_2plus_apfu == pytest.approx(0.0, abs=1e-9)  # all Fe becomes Fe3+ here


def test_apfu_matches_workbook(result, workbook_reference):
    for cation, expected in workbook_reference["apfu"].items():
        assert result.apfu[cation] == pytest.approx(expected, rel=1e-3, abs=1e-6), cation


def test_site_totals_match_workbook(result, workbook_reference):
    sites = result.site_allocation.sites
    assert not result.site_allocation.unallocated
    assert sites["T"].total == pytest.approx(workbook_reference["t_site_total"], rel=1e-3)
    ca_na_k = sites["A"].elements["Ca"] + sites["A"].elements["Na"] + sites["A"].elements["K"]
    assert ca_na_k == pytest.approx(workbook_reference["ca_na_k_total"], rel=1e-3)


def test_end_members_match_workbook(result, workbook_reference):
    em = result.end_members
    assert sum(em.values()) == pytest.approx(100.0, abs=1e-6)
    for member, expected in workbook_reference["end_members"].items():
        assert em[member] == pytest.approx(expected, rel=1e-2), member
    assert em["anorthite"] > em["albite"] > em["orthoclase"]
