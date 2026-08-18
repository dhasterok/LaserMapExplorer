"""Cross-checks the pyroxene stoichiometry pipeline against
``tests/normalization.v17.xlsx``'s 'Pyroxene' sheet -- a hand-built,
cell-by-cell recalculation of a real EPMA clinopyroxene analysis (from a
lherzolite) that solves for Fe3+ by charge balance.

Pure Python -- no PyQt/QApplication needed. ``resources/minerals/pyroxene.yaml``
uses the dedicated ``pyroxene_quad`` site-allocation method (ports MinPlotX's
``pyroxene_fe3unknown.m`` -- see ``sites.py``'s ``_allocate_pyroxene_quad``)
since MinPlotX splits Mg/Fe2+ between M1 and M2 by bulk XMg ratio, not simple
priority-fill capping.

This particular reference sample is Si-rich enough that Si (2.338 apfu)
exceeds the T site's crystallographic target (2.0) -- the workbook itself
flags this with "Note: excess tetrahedral cations" / "Note: deficiency of
octahedral cations". The workbook's own T/M1/M2 table doesn't cap T at all in
this situation (T total = raw Si = 2.338, uncapped), which is the same
uncapped-tetrahedral-site convention seen for garnet's Z site in Stage 1 --
and, per that same precedent, MinPlotX (which does cap T at 2.0, spilling
nothing further since Al isn't needed to fill it) is treated as authoritative
here, not the workbook. So T/M1/M2 site totals are checked against
expected/self-consistent values rather than the workbook directly; Fe3+/Fe2+
split, apfu, and end-members (computed from *raw* apfu, not site
occupancies, so unaffected by the T-capping disagreement) all match the
workbook tightly.

Since Si exceeds T's target here, this sample also exercises this package's
"never silently discard real composition" rule (see ``sites.py``'s module
docstring): the excess 0.338 apfu of Si ends up in ``unallocated`` rather
than vanishing the way it silently does in MinPlotX's own code (which caps
Si at 2 and never tracks the remainder).

The sheet's own oxide-total cell (B22) reads 0 -- apparently a broken
formula reference in this particular sheet -- so that one quantity isn't
cross-checked here (every other cell used below is a normal, populated
value).
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

PYROXENE_YAML_PATH = project_root / "resources" / "minerals" / "pyroxene.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"

_OXIDE_TO_CATION = {
    "SiO2": "Si", "TiO2": "Ti", "Al2O3": "Al", "Cr2O3": "Cr",
    "FeO": "Fe2", "MnO": "Mn", "MgO": "Mg", "CaO": "Ca", "Na2O": "Na", "K2O": "K",
}


@pytest.fixture(scope="module")
def pyroxene_config():
    return load_mineral_config(PYROXENE_YAML_PATH)


@pytest.fixture(scope="module")
def workbook_reference():
    ws = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)["Pyroxene"]

    analysis = {}
    expected_apfu = {}
    for row in range(11, 21):
        oxide = ws.cell(row=row, column=1).value
        if oxide is None or oxide == "Fe2O3":
            continue
        wt_pct = ws.cell(row=row, column=2).value
        if wt_pct is None:
            continue
        analysis[oxide] = float(wt_pct)
        expected_apfu[_OXIDE_TO_CATION[oxide]] = float(ws.cell(row=row, column=7).value)
    expected_apfu["Fe3"] = float(ws["G15"].value)

    return {
        "analysis": analysis,
        "apfu": expected_apfu,
        "end_members": {
            "wollastonite": float(ws["H55"].value),
            "enstatite": float(ws["H56"].value),
            "ferrosilite": float(ws["H57"].value),
        },
    }


@pytest.fixture(scope="module")
def result(pyroxene_config, workbook_reference):
    return pipeline.calculate(
        workbook_reference["analysis"], pyroxene_config,
        input_mode="wt_percent", redox_method="droop_1987",
    )


def test_fe3_charge_balance_split_matches_workbook(result, workbook_reference):
    assert result.redox.species_3plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe3"], abs=1e-6)
    assert result.redox.species_2plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe2"], rel=1e-3)


def test_apfu_matches_workbook(result, workbook_reference):
    for cation, expected in workbook_reference["apfu"].items():
        assert result.apfu[cation] == pytest.approx(expected, rel=1e-3, abs=1e-6), cation


def test_t_site_capped_and_si_excess_is_tracked_not_dropped(result):
    sites = result.site_allocation.sites
    assert sites["T"].total == pytest.approx(2.0, abs=1e-6)
    assert sites["T"].elements == pytest.approx({"Si": 2.0}, abs=1e-6)
    assert result.site_allocation.unallocated["Si"] == pytest.approx(0.3377, abs=1e-3)


def test_m1_fills_to_target_via_xmg_split_m2_takes_remainder(result):
    sites = result.site_allocation.sites
    assert sites["M1"].total == pytest.approx(1.0, abs=1e-6)
    # Al is M1-exclusive (unconditional); Mg/Fe2+ split by bulk XMg to fill
    # the rest of M1's target.
    assert sites["M1"].elements["Al"] == pytest.approx(0.14054, abs=1e-4)
    x_mg = sites["M1"].elements["Mg"] / (sites["M1"].elements["Mg"] + sites["M1"].elements["Fe2"])
    bulk_x_mg = result.apfu["Mg"] / (result.apfu["Mg"] + result.apfu["Fe2"])
    assert x_mg == pytest.approx(bulk_x_mg, rel=1e-6)
    # M2 (last site, uncapped) takes whatever Mg/Fe2+ didn't fit in M1, plus Ca.
    assert sites["M2"].elements["Ca"] == pytest.approx(0.05018, abs=1e-4)
    assert sites["M2"].elements["Mg"] + sites["M1"].elements["Mg"] == pytest.approx(result.apfu["Mg"], rel=1e-6)
    assert sites["M2"].elements["Fe2"] + sites["M1"].elements["Fe2"] == pytest.approx(result.apfu["Fe2"], rel=1e-6)


def test_end_members_match_workbook(result, workbook_reference):
    """Wo/En/Fs are computed from raw apfu (Ca/Mg/FeT), not site
    occupancies, so they're unaffected by the T-capping disagreement above --
    match the workbook tightly.
    """
    em = result.end_members
    assert sum(em.values()) == pytest.approx(100.0, abs=1e-6)
    for member, expected in workbook_reference["end_members"].items():
        assert em[member] == pytest.approx(expected, rel=1e-3), member
    assert em["enstatite"] > em["ferrosilite"] > em["wollastonite"]
