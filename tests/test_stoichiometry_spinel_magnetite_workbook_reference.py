"""Cross-checks the spinel-magnetite stoichiometry pipeline against
``tests/normalization.v17.xlsx``'s 'Spinel' sheet -- a hand-built,
cell-by-cell recalculation of a real EPMA spinel analysis (from the
Adirondacks) that solves for Fe3+ by charge balance.

Pure Python -- no PyQt/QApplication needed. ``resources/minerals/
spinel-magnetite.yaml`` (a single config covering the whole spinel group,
including magnetite -- see that file's comment) uses the dedicated
``spinel_xmg`` site-allocation method (ports MinPlotX's
``spinel_Fe3unknown.m`` -- see ``sites.py``'s ``_allocate_spinel_xmg``) since
Fe2+/Mg are equipartitioned between the A and D sites by bulk XFe/XMg ratio
rather than simple priority-fill capping.

Unlike every other mineral added in this stage, the Spinel sheet stops at
apfu/oxygen normalization -- it has no end-member section at all. So apfu and
the Fe2+/Mg D-site split (which the workbook's own numbers can still verify,
since D's Fe2+/Mg split leaves a distinct fingerprint depending on Si+Ti
content) are checked against the workbook tightly; end-members (a full port
of MinPlotX's 28-member scheme, see ``endmembers.py``'s ``_spinel_xmg``) are
only checked for internal consistency (sums to 100, dominant members are
mineralogically sane for this composition) since no authoritative reference
values exist for them here.
"""
import sys
from pathlib import Path

import openpyxl
import pytest

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))

from src.stoichiometry import pipeline
from src.stoichiometry.config import load_mineral_config

SPINEL_YAML_PATH = project_root / "resources" / "minerals" / "spinel-magnetite.yaml"
WORKBOOK_PATH = project_root / "tests" / "normalization.v17.xlsx"

_OXIDE_TO_CATION = {
    "SiO2": "Si", "TiO2": "Ti", "Al2O3": "Al", "Cr2O3": "Cr",
    "FeO": "Fe2", "MnO": "Mn", "MgO": "Mg",
}


@pytest.fixture(scope="module")
def spinel_config():
    return load_mineral_config(SPINEL_YAML_PATH)


@pytest.fixture(scope="module")
def workbook_reference():
    ws = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)["Spinel"]

    analysis = {}
    expected_apfu = {}
    for row in range(11, 19):
        oxide = ws.cell(row=row, column=1).value
        if oxide is None or oxide == "Fe2O3":
            continue
        wt_pct = ws.cell(row=row, column=2).value
        if wt_pct is None:
            continue
        analysis[oxide] = float(wt_pct)
        expected_apfu[_OXIDE_TO_CATION[oxide]] = float(ws.cell(row=row, column=7).value)
    expected_apfu["Fe3"] = float(ws["G15"].value)

    return {"analysis": analysis, "apfu": expected_apfu, "oxide_total_pct": float(ws["B20"].value)}


@pytest.fixture(scope="module")
def result(spinel_config, workbook_reference):
    return pipeline.calculate(
        workbook_reference["analysis"], spinel_config,
        input_mode="wt_percent", redox_method="droop_1987",
    )


def test_oxide_total_matches_workbook(result, workbook_reference):
    assert result.oxide_total_pct == pytest.approx(workbook_reference["oxide_total_pct"], rel=1e-3)


def test_fe3_charge_balance_split_matches_workbook(result, workbook_reference):
    assert result.redox.species_3plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe3"], rel=1e-3)
    assert result.redox.species_2plus_apfu == pytest.approx(workbook_reference["apfu"]["Fe2"], rel=1e-3)


def test_apfu_matches_workbook(result, workbook_reference):
    for cation, expected in workbook_reference["apfu"].items():
        assert result.apfu[cation] == pytest.approx(expected, rel=1e-3, abs=1e-6), cation


def test_d_and_a_site_totals_and_fe2_mg_equipartition(result):
    sites = result.site_allocation.sites
    assert not result.site_allocation.unallocated
    assert sites["D"].total + sites["A"].total == pytest.approx(3.0, abs=1e-6)
    # D/A individually deviate from their nominal 2/1 targets by exactly how
    # much Ti (A-exclusive) spills in, since nothing compensates for it --
    # expected, not a bug (matches MinPlotX's own unbounded A/D split).
    assert sites["D"].total == pytest.approx(2.0, abs=0.01)
    assert sites["A"].total == pytest.approx(1.0, abs=0.01)
    # Al/Fe3+ (D-exclusive) are unconditional and dominate D for this
    # Al-rich sample.
    assert sites["D"].elements["Al"] == pytest.approx(1.921, abs=1e-3)
    assert sites["D"].elements["Fe3"] == pytest.approx(0.0741, abs=1e-3)
    # Only a small fraction of Fe2+/Mg spills into D, scaled by Ti (the only
    # A-exclusive cation present in this sample, Si=0) relative to bulk Mg+Fe2+.
    si_ti = result.apfu.get("Si", 0.0) + result.apfu["Ti"]
    assert (sites["D"].elements.get("Fe2", 0.0) + sites["D"].elements.get("Mg", 0.0)) == pytest.approx(si_ti, rel=1e-6)


def test_end_members_are_internally_consistent(result, spinel_config):
    """No workbook reference exists for spinel end-members (see module
    docstring) -- sanity-check internal consistency and that the dominant
    members are mineralogically sane for this Al-rich, Fe/Mg-bearing,
    essentially Si/Cr/V-free composition (should be spinel s.s. +
    hercynite, not any Cr/V/Si/Ti-bearing member). Only the tiered
    `members` (not the separate `ratios`, mixed into the same dict -- see
    config.py's EndMemberConfig.ratios docstring) sum to 100.
    """
    em = result.end_members
    members_sum = sum(em[m] for m in spinel_config.end_members.members)
    assert members_sum == pytest.approx(100.0, abs=1e-6)
    assert all(em[m] >= 0 for m in spinel_config.end_members.members)
    assert em["spinel"] > em["hercynite"] > 0
    assert em["spinel"] + em["hercynite"] > 90  # overwhelmingly Mg-Al with minor Fe2+
    for cr_or_v_member in ("chromite", "magnesiochromite", "coulsonite", "magnesiocoulsonite"):
        assert em[cr_or_v_member] == pytest.approx(0.0, abs=1e-6)
